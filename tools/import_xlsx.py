#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Импорт телефонного справочника из .xlsx в data/rukovoditeli.json.

    python3 tools/import_xlsx.py путь/к/справочнику.xlsx

Из справочника берутся только руководящие должности — главы, начальники,
председатели, директора, руководители и их заместители. Рядовые сотрудники
в data/rukovoditeli.json не попадают: раздел «Визитница» — о руководителях.

Ожидаемые колонки листа, начиная со строки 5:
    A Подразделение · B Отдел · C Должность · D ФИО · E Кабинет · F Телефон · G E-mail
Пустая ячейка A означает «то же подразделение, что строкой выше».

Требуется openpyxl:  pip install openpyxl
"""
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / 'data' / 'rukovoditeli.json'

# Должности, которые считаются руководящими.
KEYWORDS = ('глава', 'главы', 'начальник', 'председател', 'директор',
            'руководител', 'управляющий делами', 'заведующ')
# «Советник главы» — не руководитель подразделения, хотя и содержит «главы».
EXCLUDE = ('советник',)
DEPUTY = re.compile(r'^\s*(заместител|первый заместител|и\.\s*о\.)', re.I)


def is_leader(post):
    p = post.lower()
    if any(x in p for x in EXCLUDE):
        return False
    return any(k in p for k in KEYWORDS)


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__.strip())
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit('Файл не найден: %s' % src)

    try:
        import openpyxl
    except ImportError:
        sys.exit('Нужен openpyxl:  pip install openpyxl')

    ws = openpyxl.load_workbook(src, data_only=True).active

    people, group = [], ''
    for row in ws.iter_rows(min_row=5, values_only=True):
        cells = [('' if c is None else str(c).strip()) for c in row]
        if not any(cells):
            continue
        if cells[0]:
            group = cells[0]
        people.append(OrderedDict(
            подразделение=group, отдел=cells[1], должность=cells[2],
            фио=cells[3], кабинет=cells[4], телефон=cells[5], почта=cells[6]))

    # Штатная численность считается по всем строкам, в файл идут только руководители.
    счёт = OrderedDict()
    for p in people:
        ключ = (p['подразделение'], p['отдел'])
        счёт[ключ] = счёт.get(ключ, 0) + 1
    штат = [OrderedDict(подразделение=k[0], отдел=k[1], человек=v)
            for k, v in счёт.items()]

    leaders = [p for p in people if is_leader(p['должность'])]
    for p in leaders:
        p['заместитель'] = bool(DEPUTY.match(p['должность']))

    data = OrderedDict(
        источник=str(ws['A1'].value or src.name).strip(),
        актуально=str(ws['A2'].value or '').strip(),
        численность=штат,
        руководители=leaders,
    )
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print('Всего строк справочника: %d' % len(people))
    print('Руководителей отобрано:  %d' % len(leaders))
    print('Записано: %s' % OUT.relative_to(ROOT))


if __name__ == '__main__':
    main()
